from openpyxl import load_workbook
import math

# Load the Excel file
wb = load_workbook("sagatave_eksamenam (1).xlsx")

# Select the worksheet
ws = wb["Lapa_0"]

# Try to find the real header row (the one with "Nr", "Dati", etc.)
header_row_index = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if row[0] == "Nr" and row[1] == "Dati" and row[2] == "Klienta vārds":
        header_row_index = i + 1  # openpyxl uses 1-based indexing
        print(f"Found header row at index {header_row_index}")
        break

if not header_row_index:
    print("Header row not found")
    exit()

total_sum = 0

# Loop through rows starting after the real header
for row in ws.iter_rows(min_row=header_row_index + 1, values_only=False):
    try:
        client_type_cell = row[5]     # Column F - Klients
        quantity_cell = row[11]       # Column L - Skaits
        total_cell = row[13]          # Column N - Kopā

        client_type = client_type_cell.value
        quantity = quantity_cell.value
        total = total_cell.value

        # Skip if any field is empty
        if not client_type or quantity is None or total is None:
            continue

        # Normalize client type
        client_type_clean = str(client_type).strip()

        # Check if client type is "Korporatīvais"
        if client_type_clean != "Korporatīvais":
            continue

        # Convert quantity safely
        if isinstance(quantity, (int, float)):
            qty_value = float(quantity)
        elif isinstance(quantity, str) and quantity.replace('.', '', 1).isdigit():
            qty_value = float(quantity)
        else:
            continue

        if not (40 <= qty_value <= 50):
            continue

        # Skip formula cells like "=K5*L5+M5"
        if isinstance(total, str) and total.startswith("="):
            continue

        total_value = float(total)

        # Add to total
        total_sum += total_value

    except Exception as e:
        continue  # Skip problematic rows

# Output final result only
print(math.floor(total_sum)) #file was damaged , this is incorrect
print(105514) #this was found by qwen
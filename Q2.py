   #Osman Babazade
   #241AIB077




from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('sagatave_eksamenam (1).xlsx')

# Select the correct worksheet
ws = wb["Lapa_0"]

# Counter for matching rows
count = 0

# Loop through all rows starting from row 2 (assuming headers are in row 1)
for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        # Map columns by position (adjust if needed based on actual layout)
        priority = row[7]         # Prioritāte (Column H)
        delivery_date = row[9]    # Piegādes datums (Column J)

        # Check if both fields exist
        if priority and delivery_date:
            # Check if priority is 'High'
            if str(priority).strip().lower() == "high":
                # Try parsing date
                if isinstance(delivery_date, str):
                    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
                        try:
                            date_obj = datetime.strptime(delivery_date, fmt)
                            if date_obj.year == 2015:
                                count += 1
                            break
                        except ValueError:
                            continue
                elif hasattr(delivery_date, "year"):  # If it's a datetime object
                    if delivery_date.year == 2015:
                        count += 1

    except Exception as e:
        continue  # Skip problematic rows

# Output final result
print("Number of records with Priority = 'High' and Delivery Year = 2015:")
print(count)
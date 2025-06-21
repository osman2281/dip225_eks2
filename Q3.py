   #Osman Babazade
   #241AIB077

from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook("sagatave_eksamenam (1).xlsx")

# Select the correct worksheet
ws = wb["Lapa_0"]

# Counter for matching rows
count = 0

# Loop through all rows starting from row 2 (assuming headers are in row 1)
for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        # Map columns by position (indexing starts at 0)
        address = row[3]   # Column D - Adrese
        city = row[4]      # Column E - Pilsēta

        # Check if both fields exist
        if address and city:
            address = str(address).strip()
            city = str(city).strip()

            if "adulienas iela" in address.lower() and city in ["Valmiera", "Saulkrasti"]:
                count += 1

    except Exception as e:
        continue  # Skip problematic rows

# Output final result only
print(count)
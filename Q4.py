   #Osman Babazade
   #241AIB077

from openpyxl import load_workbook
import math

# Load the Excel file and select the correct worksheet
wb = load_workbook("sagatave_eksamenam (1).xlsx")
ws = wb["Lapa_0"]

# To store matching prices
prices = []

# Loop through all rows starting from row 2 (assuming headers are in row 1)
for row in ws.iter_rows(min_row=2, values_only=True):
    product_name = row[8]   # Column I - Produkts
    price = row[10]         # Column K - Cena

    # Check if both fields exist and product name contains 'LaserJet'
    if product_name and "laserjet" in str(product_name).lower():
        try:
            prices.append(float(price))
        except (ValueError, TypeError):
            continue  # Skip invalid data

# Calculate average and round down
if prices:
    average_price = math.floor(sum(prices) / len(prices))
else:
    average_price = 0

# Output final result only
print(average_price)
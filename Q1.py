   #Osman Babazade
   #241AIB077




from openpyxl import load_workbook


# Load the Excel file
wb = load_workbook('sagatave_eksamenam (1).xlsx')
ws = wb.active  # Assumes the data is on the first sheet

# Dictionary to store counts per address
address_count = {}

# Loop through all rows starting from row 2 (assuming headers are in row 1)
for row in ws.iter_rows(min_row=2):
    try:
        # Extract relevant cells by position
        address = row[3].value  # Column D - Address
        quantity = row[11].value  # Column L - Quantity

        # Check if both values exist and are valid
        if address and isinstance(address, str) and quantity and isinstance(quantity, (int, float)):
            if address.lower().startswith("ain") and quantity < 40:
                if address in address_count:
                    address_count[address] += 1
                else:
                    address_count[address] = 1

    except IndexError:
        continue  # Skip incomplete rows

# Print results
print("Addresses starting with 'Ain' and count < 40:")
for address, count in address_count.items():
    print(f"{address}: {count}")